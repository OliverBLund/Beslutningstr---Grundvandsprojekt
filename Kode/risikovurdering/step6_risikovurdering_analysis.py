"""
Step 6: Final Comprehensive Analysis
=====================================

BASELINE COMPARISON - IMPORTANT:
This analysis compares against Step 5b (compound-specific) NOT Step 5a (general 500m).

Core Scenario (Step 5b):
- Sites WITH substance data using compound-specific distance thresholds
- Source: step5_compound_detailed_combinations.csv (all site-GVFK-substance combinations)
- Result: ~1,740 unique sites spanning 240 GVFKs
- These sites qualified under literature-based variable thresholds per compound category
- Note: 240 GVFKs because multi-GVFK approach preserves all site-GVFK associations

Expanded Scenario (Step 5b⁺):
- Core + branch-only sites (<=500m, no losseplads keywords)
- Branch-only source: step5_unknown_substance_sites.csv (sites WITHOUT substance data)
- Filtering: Final_Distance_m <= 500 AND excludes losseplads keywords (see lines 129-139)
- Result: Core (240 GVFKs) + New GVFKs = total expanded GVFKs
- The "new" GVFKs have ONLY branch-only sites, no substance sites

KEY DISTINCTION vs step5_branch_analysis.py:
- step5_branch_analysis.py compares branch-only sites against Step 5a (general 500m assessment)
  which includes ALL substance sites within 500m (~300+ GVFKs), resulting in additional GVFKs
- THIS analysis (step6) compares against Step 5b (compound-specific, 240 GVFKs)
- GVFK counts are calculated dynamically from detailed combinations to preserve multi-GVFK associations
- See step5_branch_analysis.py lines 453-541 (_analyze_generel_risiko_impact) for comparison

Analysis outputs:
Phase 1: Data loading and GVFK categorization (lines 42-292)
Phase 2: GVFK progression metrics and visualizations (lines 294-656)
Phase 3: Three-way branch/activity comparison 
Phase 4: Geographic visualizations 
"""

import pandas as pd
import geopandas as gpd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import os
import sys
from pathlib import Path
from collections import Counter

if __package__ is None or __package__ == "":
    sys.path.append(str(Path(__file__).resolve().parents[1]))

from config import get_output_path, get_visualization_path, GRUNDVAND_PATH, WORKFLOW_SETTINGS, GVFK_AREA_VOLUME_PATH, DATA_DIR

# Professional styling with larger fonts for Word documents (50% increase)
plt.rcParams.update({
    'font.family': ['Arial', 'DejaVu Sans', 'sans-serif'],
    'font.size': 14,           # Default (was ~10)
    'axes.titlesize': 20,       # Titles (was 12-15)
    'axes.labelsize': 16,       # Axis labels (was 11-12)
    'xtick.labelsize': 14,      # X-tick labels (was 9-10)
    'ytick.labelsize': 14,      # Y-tick labels (was 9-10)
    'legend.fontsize': 14,      # Legend (was 9-10)
    'figure.titlesize': 22,     # Figure title (was 14-15)
    'figure.dpi': 100,
    'savefig.dpi': 300,
    'savefig.bbox': 'tight'
})

COLORS = {
    'core': '#1E88E5',           # Blue for core scenario
    'expanded': '#FFC107',        # Amber for expanded
    'new_gvfk': '#FF5722',       # Red-orange for new GVFKs
    'shared': '#43A047',          # Green for shared
    'neutral': '#757575'          # Gray
}


# ============================================================================
# FILE CONFIGURATION - ALL INPUT/OUTPUT FILES IN ONE PLACE
# ============================================================================
"""
Centralized file paths configuration. All files used in Step 6 are defined here.
This makes it easy to see what files are needed and change paths if necessary.
"""

# INPUT FILES - All upstream outputs and data files
INPUT_FILES = {
    # Step 5 outputs
    'step5_compound_detailed': get_output_path('step5_compound_detailed_combinations'),  # Core scenario
    'step5_parked_sites': get_output_path('step5_unknown_substance_sites'),  # Parked sites

    # Step 4 output (for site geometries)
    'step4_sites_with_geometry': get_output_path('unique_lokalitet_distances_shp'),

    # Reference data
    'gvfk_area_volume': GVFK_AREA_VOLUME_PATH,

    # Shapefiles for visualization
    'grundvand_shapefile': GRUNDVAND_PATH,
    'step2_gvfk_shapefile': get_output_path('step2_river_gvfk'),
    'step3_gvfk_shapefile': get_output_path('step3_gvfk_polygons'),
    'step3_sites_shapefile': get_output_path('step3_v1v2_sites'),
    'denmark_regions_shapefile': DATA_DIR / 'regionsinddeling',  # Optional backdrop for maps
}

# OUTPUT DIRECTORY - All visualizations and tables go here
OUTPUT_DIR = get_visualization_path('step6')


# ============================================================================
# PHASE 1: DATA LOADING AND PREPARATION
# ============================================================================

def load_gvfk_area_volume():
    """
    Load GVFK area and volume data from CSV.

    Returns:
        dict: {gvfk_name: {'area_km2': float, 'volume_m³': float}}
    """
    print("[PHASE 1] Loading GVFK area/volume data...")

    file_path = INPUT_FILES['gvfk_area_volume']
    if not file_path.exists():
        print(f"Warning: Area/volume file not found at {file_path}")
        return {}

    df = pd.read_csv(file_path, sep=';', decimal=',', encoding='utf-8')

    gvfk_data = {}
    for _, row in df.iterrows():
        gvfk_name = row['GVFK']
        try:
            area = float(row['Areal [km2]'])
            volume = float(row['Volumen'])
            gvfk_data[gvfk_name] = {
                'area_km2': area,
                'volume_m³': volume
            }
        except (ValueError, KeyError):
            continue

    print(f"  Loaded area/volume data for {len(gvfk_data)} GVFKs")
    return gvfk_data

def load_substance_sites():
    """
    Load Step 5b compound-specific detailed combinations (Core scenario).

    Loads all site-GVFK-substance combinations to preserve GVFK associations,
    then deduplicates by site for site-level analyses.

    Returns:
        tuple: (detailed_combinations DataFrame, deduplicated_sites DataFrame)
            - detailed_combinations: All site-GVFK-substance combinations (for GVFK counting)
            - deduplicated_sites: One row per site (for site-level analyses)
    """
    print("\n[PHASE 1] Loading substance sites (Core scenario)...")

    # Load detailed combinations to preserve all GVFK associations
    detailed_path = INPUT_FILES['step5_compound_detailed']
    if not os.path.exists(detailed_path):
        raise FileNotFoundError(f"Compound-specific detailed combinations not found: {detailed_path}")

    detailed_combinations = pd.read_csv(detailed_path)

    # Standardize column names for compatibility with step6
    if 'GVFK' in detailed_combinations.columns and 'Closest_GVFK' not in detailed_combinations.columns:
        detailed_combinations = detailed_combinations.rename(columns={'GVFK': 'Closest_GVFK'})

    if 'Distance_to_River_m' in detailed_combinations.columns and 'Final_Distance_m' not in detailed_combinations.columns:
        detailed_combinations = detailed_combinations.rename(columns={'Distance_to_River_m': 'Final_Distance_m'})

    # Calculate statistics from detailed combinations (correct GVFK count)
    total_combinations = len(detailed_combinations)
    n_unique_sites = detailed_combinations['Lokalitet_ID'].nunique()
    n_gvfks = detailed_combinations['Closest_GVFK'].dropna().nunique()

    print(f"  Loaded {total_combinations:,} site-GVFK-substance combinations")
    print(f"  → {n_unique_sites:,} unique sites across {n_gvfks} GVFKs")

    # Deduplicate by site (keep first occurrence) for site-level analyses
    substance_sites = detailed_combinations.drop_duplicates(subset=['Lokalitet_ID'], keep='first').copy()

    return detailed_combinations, substance_sites


def load_and_filter_branch_sites():
    """
    Load and filter branch-only sites (<=500m, no losseplads).

    Returns:
        pd.DataFrame: Filtered branch-only sites
    """
    print("\n[PHASE 1] Loading and filtering branch-only sites...")

    parked_path = INPUT_FILES['step5_parked_sites']
    if not os.path.exists(parked_path):
        raise FileNotFoundError(f"Parked sites not found: {parked_path}")

    unknown_sites = pd.read_csv(parked_path)

    # Standardize column names for compatibility with step6
    if 'GVFK' in unknown_sites.columns and 'Closest_GVFK' not in unknown_sites.columns:
        unknown_sites = unknown_sites.rename(columns={'GVFK': 'Closest_GVFK'})

    if 'Distance_to_River_m' in unknown_sites.columns and 'Final_Distance_m' not in unknown_sites.columns:
        unknown_sites = unknown_sites.rename(columns={'Distance_to_River_m': 'Final_Distance_m'})

    print(f"  Total unknown substance sites: {len(unknown_sites):,}")

    # Filter 1: Distance <=500m
    branch_sites = unknown_sites[unknown_sites['Final_Distance_m'] <= 500].copy()
    print(f"  After <=500m filter: {len(branch_sites):,} sites")

    # Filter 2: Exclude losseplads keywords
    def has_losseplads(text):
        if pd.isna(text):
            return False
        text_lower = str(text).lower()
        keywords = ['losseplads', 'affald', 'depon', 'fyldplads', 'skraldeplads']
        return any(keyword in text_lower for keyword in keywords)

    branch_sites = branch_sites[
        ~(branch_sites['Lokalitetensbranche'].apply(has_losseplads) |
          branch_sites['Lokalitetensaktivitet'].apply(has_losseplads))
    ].copy()

    n_sites = len(branch_sites)
    n_gvfks = branch_sites['Closest_GVFK'].dropna().nunique()

    print(f"  After losseplads exclusion: {n_sites:,} sites")
    print(f"  Branch-only sites span {n_gvfks} GVFKs")

    return branch_sites


def categorize_gvfks(substance_detailed, branch_sites):
    """
    Categorize GVFKs into shared, new, and substance-only groups.

    Args:
        substance_detailed: Core scenario detailed combinations (ALL site-GVFK-substance combinations)
        branch_sites: Branch-only sites

    Returns:
        dict: {
            'core_gvfks': set,
            'branch_gvfks': set,
            'shared_gvfks': set,
            'new_gvfks': set,
            'substance_only_gvfks': set,
            'expanded_gvfks': set
        }
    """
    print("\n[PHASE 1] Categorizing GVFKs...")

    # Use detailed combinations to get ALL GVFKs (preserves multi-GVFK associations)
    core_gvfks = set(substance_detailed['Closest_GVFK'].dropna().unique())
    branch_gvfks = set(branch_sites['Closest_GVFK'].dropna().unique())

    shared_gvfks = core_gvfks & branch_gvfks
    new_gvfks = branch_gvfks - core_gvfks
    substance_only_gvfks = core_gvfks - branch_gvfks
    expanded_gvfks = core_gvfks | branch_gvfks

    print(f"  Core GVFKs (substance sites): {len(core_gvfks)}")
    print(f"  Branch GVFKs (total): {len(branch_gvfks)}")
    print(f"  Shared GVFKs (both types): {len(shared_gvfks)}")
    print(f"  New GVFKs (branch-only): {len(new_gvfks)}")
    print(f"  Substance-only GVFKs: {len(substance_only_gvfks)}")
    print(f"  Expanded GVFKs (total): {len(expanded_gvfks)}")

    # Verify math
    assert len(shared_gvfks) + len(substance_only_gvfks) == len(core_gvfks)
    assert len(shared_gvfks) + len(new_gvfks) == len(branch_gvfks)

    return {
        'core_gvfks': core_gvfks,
        'branch_gvfks': branch_gvfks,
        'shared_gvfks': shared_gvfks,
        'new_gvfks': new_gvfks,
        'substance_only_gvfks': substance_only_gvfks,
        'expanded_gvfks': expanded_gvfks
    }


def filter_sites_by_gvfk_category(branch_sites, new_gvfks):
    """
    Filter branch-only sites to those in the "new" GVFKs.

    Args:
        branch_sites: All branch-only sites
        new_gvfks: Set of new GVFK names

    Returns:
        pd.DataFrame: Sites in new GVFKs only
    """
    print("\n[PHASE 1] Filtering sites in new GVFKs...")

    sites_in_new_gvfks = branch_sites[
        branch_sites['Closest_GVFK'].isin(new_gvfks)
    ].copy()

    print(f"  Sites in new {len(new_gvfks)} GVFKs: {len(sites_in_new_gvfks):,}")

    return sites_in_new_gvfks


def load_gvfk_shapefiles():
    """
    Load GVFK shapefiles for geographic analysis.

    Returns:
        dict: {'all_dk': GeoDataFrame, 'step2': GeoDataFrame, 'step3': GeoDataFrame}
    """
    print("\n[PHASE 1] Loading GVFK shapefiles...")

    shapefiles = {}

    # All Denmark
    all_dk_path = INPUT_FILES['grundvand_shapefile']
    if os.path.exists(all_dk_path):
        shapefiles['all_dk'] = gpd.read_file(all_dk_path)
        print(f"  All Denmark: {len(shapefiles['all_dk'])} GVFKs")

    # Step 2
    step2_path = INPUT_FILES['step2_gvfk_shapefile']
    if os.path.exists(step2_path):
        shapefiles['step2'] = gpd.read_file(step2_path)
        print(f"  Step 2 (river contact): {len(shapefiles['step2'])} GVFKs")

    # Step 3
    step3_path = INPUT_FILES['step3_gvfk_shapefile']
    if os.path.exists(step3_path):
        shapefiles['step3'] = gpd.read_file(step3_path)
        print(f"  Step 3 (V1/V2 sites): {len(shapefiles['step3'])} GVFKs")

    return shapefiles


def run_phase1():
    """
    Execute Phase 1: Load and prepare all data.

    Returns:
        dict: All loaded data and categorizations
    """
    print("\n" + "=" * 60)
    print("PHASE 1: DATA LOADING AND PREPARATION")
    print("=" * 60)

    # Load data
    gvfk_area_volume = load_gvfk_area_volume()
    substance_detailed, substance_sites = load_substance_sites()
    branch_sites = load_and_filter_branch_sites()

    # Categorize GVFKs (use detailed combinations to preserve all GVFK associations)
    gvfk_categories = categorize_gvfks(substance_detailed, branch_sites)

    # Filter sites in new GVFKs
    sites_in_new_gvfks = filter_sites_by_gvfk_category(
        branch_sites,
        gvfk_categories['new_gvfks']
    )

    # Load shapefiles
    shapefiles = load_gvfk_shapefiles()

    print("\n" + "=" * 60)
    print("[OK] PHASE 1 COMPLETE")
    print("=" * 60)

    return {
        'gvfk_area_volume': gvfk_area_volume,
        'substance_detailed': substance_detailed,  # For accurate GVFK counting
        'substance_sites': substance_sites,
        'branch_sites': branch_sites,
        'sites_in_new_gvfks': sites_in_new_gvfks,
        'gvfk_categories': gvfk_categories,
        'shapefiles': shapefiles
    }


# ============================================================================
# PHASE 2: GVFK PROGRESSION METRICS
# ============================================================================

def calculate_gvfk_metrics(gvfk_set, gvfk_area_volume):
    """
    Calculate aggregate metrics for a set of GVFKs.

    Args:
        gvfk_set: Set of GVFK names
        gvfk_area_volume: Dictionary with area/volume data per GVFK

    Returns:
        dict: {'count': int, 'area_km2': float, 'volume_m³': float}
    """
    count = len(gvfk_set)
    total_area = 0.0
    total_volume = 0.0

    for gvfk in gvfk_set:
        if gvfk in gvfk_area_volume:
            total_area += gvfk_area_volume[gvfk]['area_km2']
            total_volume += gvfk_area_volume[gvfk]['volume_m³']

    return {
        'count': count,
        'area_km2': total_area,
        'volume_m³': total_volume
    }


def create_gvfk_count_progression_chart(shapefiles, gvfk_categories, output_dir):
    """
    Create bar chart showing GVFK count progression through workflow.

    Args:
        shapefiles: Dict with 'all_dk', 'step2', 'step3' GeoDataFrames
        gvfk_categories: Dict with 'core_gvfks', 'expanded_gvfks'
        output_dir: Output directory path
    """
    print("\n[PHASE 2] Creating GVFK count progression chart...")

    # Calculate counts
    steps = []
    counts = []

    if 'all_dk' in shapefiles:
        steps.append('Trin 1:\nAlle Danmark')
        counts.append(len(shapefiles['all_dk']))

    if 'step2' in shapefiles:
        steps.append('Trin 2:\nVandløbskontakt')
        counts.append(len(shapefiles['step2']))

    if 'step3' in shapefiles:
        steps.append('Trin 3:\nV1/V2 Lokaliteter')
        counts.append(len(shapefiles['step3']))

    steps.append('Trin 5b:\nKerne\n(Substans)')
    counts.append(len(gvfk_categories['core_gvfks']))

    steps.append('Trin 5b⁺:\nUdvidet\n(+Branche)')
    counts.append(len(gvfk_categories['expanded_gvfks']))

    # Create figure
    fig, ax = plt.subplots(figsize=(10, 6))

    # Create bars with different colors
    colors = ['#90CAF9', '#64B5F6', '#42A5F5', COLORS['core'], COLORS['expanded']]
    bars = ax.bar(steps, counts, color=colors, edgecolor='black', linewidth=0.5)

    # Add value labels on bars
    for bar, count in zip(bars, counts):
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{count:,}',
                ha='center', va='bottom', fontweight='bold', fontsize=16)

    # Formatting
    ax.set_ylabel('Antal GVFKs', fontweight='bold')
    ax.set_title('GVFK Progression Gennem Workflow', fontweight='bold', fontsize=20, pad=15)
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    ax.set_axisbelow(True)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'gvfk_count_progression.png'), dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  Saved: gvfk_count_progression.png")


def create_area_volume_progression_chart(shapefiles, gvfk_categories, gvfk_area_volume, output_dir):
    """Create dual-axis chart showing area/volume progression."""
    print("\n[PHASE 2] Creating area/volume progression chart...")

    labels = []
    area_core = []
    area_branch = []
    volume_core = []
    volume_branch = []
    baseline_area = None
    baseline_volume = None

    def append_stage(label, gvfk_names):
        metrics = calculate_gvfk_metrics(gvfk_names, gvfk_area_volume)
        labels.append(label)
        area_core.append(metrics['area_km2'])
        area_branch.append(0.0)
        volume_core.append(metrics['volume_m³'])
        volume_branch.append(0.0)
        return metrics

    if 'all_dk' in shapefiles:
        metrics = append_stage('Trin 1:\nAlle Danmark', set(shapefiles['all_dk']['Navn'].unique()))
        baseline_area = metrics['area_km2']
        baseline_volume = metrics['volume_m³']

    if 'step2' in shapefiles:
        append_stage('Trin 2:\nVandløbskontakt', set(shapefiles['step2']['Navn'].unique()))

    if 'step3' in shapefiles:
        append_stage('Trin 3:\nV1/V2 Lokaliteter', set(shapefiles['step3']['Navn'].unique()))

    core_metrics = append_stage('Trin 5b:\nKerne (Substans)', gvfk_categories['core_gvfks'])
    core_area = core_metrics['area_km2']
    core_volume = core_metrics['volume_m³']

    new_metrics = calculate_gvfk_metrics(gvfk_categories['new_gvfks'], gvfk_area_volume)
    labels.append(f"Trin 5b⁺:\nUdvidet (+{len(gvfk_categories['new_gvfks'])} GVFKs)")
    area_core.append(core_area)
    area_branch.append(new_metrics['area_km2'])
    volume_core.append(core_volume)
    volume_branch.append(new_metrics['volume_m³'])

    x = np.arange(len(labels))
    width = 0.35

    area_core_arr = np.array(area_core)
    area_branch_arr = np.array(area_branch)
    volume_core_arr = np.array(volume_core)
    volume_branch_arr = np.array(volume_branch)

    fig, ax1 = plt.subplots(figsize=(12, 7))

    ax1.bar(x - width/2, area_core_arr, width, color='#4CAF50', alpha=0.85,
            edgecolor='black', linewidth=0.6, label='Areal (km²) – kerne del')
    ax1.bar(x - width/2, area_branch_arr, width, bottom=area_core_arr,
            color='#FFC107', alpha=0.8, edgecolor='black', linewidth=0.6,
            label='Areal (km²) – branche-kun tilføjelse')

    ax1.set_ylabel('Areal (km²)', fontweight='bold', color='#4CAF50')
    ax1.tick_params(axis='y', labelcolor='#4CAF50')
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, fontsize=14)

    # Add y-axis padding (15% extra space above max)
    ax1.set_ylim(0, max(area_core_arr + area_branch_arr) * 1.15)

    ax2 = ax1.twinx()
    ax2.bar(x + width/2, volume_core_arr, width, color='#1E88E5', alpha=0.85,
            edgecolor='black', linewidth=0.6, label='Volumen (m³) – kerne del')
    ax2.bar(x + width/2, volume_branch_arr, width, bottom=volume_core_arr,
            color='#FB8C00', alpha=0.8, edgecolor='black', linewidth=0.6,
            label='Volumen (m³) – branche-kun tilføjelse')

    ax2.set_ylabel('Volumen (m³)', fontweight='bold', color='#1E88E5')
    ax2.tick_params(axis='y', labelcolor='#1E88E5')

    # Add y-axis padding (15% extra space above max)
    ax2.set_ylim(0, max(volume_core_arr + volume_branch_arr) * 1.15)

    # Add percentage labels inside bars
    if baseline_area and baseline_area > 0:
        for i in range(len(x)):
            total_area = area_core_arr[i] + area_branch_arr[i]
            area_pct = (total_area / baseline_area) * 100

            # For Step 5b+ (last bar), show total and additional percentage points
            if i == len(x) - 1 and area_branch_arr[i] > 0:
                core_pct = (area_core_arr[i] / baseline_area) * 100
                additional_pct = area_pct - core_pct  # Percentage points added
                ax1.text(x[i] - width/2, total_area/2, f'{area_pct:.0f}%\n(+{additional_pct:.1f}%)',
                         ha='center', va='center', fontweight='bold', color='white', fontsize=13)
            else:
                ax1.text(x[i] - width/2, total_area/2, f'{area_pct:.0f}%',
                         ha='center', va='center', fontweight='bold', color='white', fontsize=14)

    if baseline_volume and baseline_volume > 0:
        for i in range(len(x)):
            total_volume = volume_core_arr[i] + volume_branch_arr[i]
            volume_pct = (total_volume / baseline_volume) * 100

            # For Step 5b+ (last bar), show total and additional percentage points
            if i == len(x) - 1 and volume_branch_arr[i] > 0:
                core_pct = (volume_core_arr[i] / baseline_volume) * 100
                additional_pct = volume_pct - core_pct  # Percentage points added
                ax2.text(x[i] + width/2, total_volume/2, f'{volume_pct:.0f}%\n(+{additional_pct:.1f}%)',
                         ha='center', va='center', fontweight='bold', color='white', fontsize=13)
            else:
                ax2.text(x[i] + width/2, total_volume/2, f'{volume_pct:.0f}%',
                         ha='center', va='center', fontweight='bold', color='white', fontsize=14)

    fig.suptitle('GVFK Areal og Volumen Progression', fontsize=22, fontweight='bold')
    ax1.grid(axis='y', alpha=0.3)
    ax1.margins(x=0.05)

    # Move legend to upper right inside the plot
    handles1, labels1 = ax1.get_legend_handles_labels()
    handles2, labels2 = ax2.get_legend_handles_labels()
    by_label = dict(zip(labels1 + labels2, handles1 + handles2))
    ax1.legend(by_label.values(), by_label.keys(), loc='upper right', fontsize=13, framealpha=0.95)

    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'gvfk_area_volume_progression.png'), dpi=300, bbox_inches='tight')
    plt.close()
    print('  Saved: gvfk_area_volume_progression.png')

    if baseline_area and baseline_area > 0:
        expanded_area = core_area + area_branch_arr[-1]
        expanded_volume = core_volume + volume_branch_arr[-1]
        print('  Area/volume retention relative to baseline:')
        print(f"    Step 5b (core): {core_area:,.0f} km² ({core_area / baseline_area * 100:.1f}% of baseline area)")
        print(f"    Step 5b⁺ (expanded): {expanded_area:,.0f} km² ({expanded_area / baseline_area * 100:.1f}% of baseline area)")
        if baseline_volume and baseline_volume > 0:
            print(f"    Step 5b (core): {core_volume:,.2e} m³ ({core_volume / baseline_volume * 100:.1f}% of baseline volume)")
            print(f"    Step 5b⁺ (expanded): {expanded_volume:,.2e} m³ ({expanded_volume / baseline_volume * 100:.1f}% of baseline volume)")
        if core_area > 0:
            print(f"    Additional area from branch-only GVFKs: {area_branch_arr[-1]:,.0f} km² (+{area_branch_arr[-1] / core_area * 100:.1f}% vs core)")
        if core_volume > 0:
            print(f"    Additional volume from branch-only GVFKs: {volume_branch_arr[-1]:.2e} m³ (+{volume_branch_arr[-1] / core_volume * 100:.1f}% vs core)")



def create_progression_table(shapefiles, gvfk_categories, gvfk_area_volume, output_dir):
    """
    Create detailed progression table with percentages.

    Args:
        shapefiles: Dict with GeoDataFrames
        gvfk_categories: Dict with GVFK sets
        gvfk_area_volume: Dict with area/volume data
        output_dir: Output directory path

    Returns:
        pd.DataFrame: Progression table
    """
    print("\n[PHASE 2] Creating progression table...")

    rows = []
    baseline_count = baseline_area = baseline_volume = None

    # Step 1: All Denmark (baseline)
    if 'all_dk' in shapefiles:
        gvfk_set = set(shapefiles['all_dk']['Navn'].unique())
        metrics = calculate_gvfk_metrics(gvfk_set, gvfk_area_volume)
        rows.append({
            'Step': 'Step 1: All Denmark',
            'GVFK Count': metrics['count'],
            'Area (km²)': metrics['area_km2'],
            'Volume (m³)': metrics['volume_m³']
        })
        baseline_count = metrics['count']
        baseline_area = metrics['area_km2']
        baseline_volume = metrics['volume_m³']

    # Step 2: River Contact
    if 'step2' in shapefiles:
        gvfk_set = set(shapefiles['step2']['Navn'].unique())
        metrics = calculate_gvfk_metrics(gvfk_set, gvfk_area_volume)
        rows.append({
            'Step': 'Step 2: River Contact',
            'GVFK Count': metrics['count'],
            'Area (km²)': metrics['area_km2'],
            'Volume (m³)': metrics['volume_m³']
        })

    # Step 3: V1/V2 Sites
    if 'step3' in shapefiles:
        gvfk_set = set(shapefiles['step3']['Navn'].unique())
        metrics = calculate_gvfk_metrics(gvfk_set, gvfk_area_volume)
        rows.append({
            'Step': 'Step 3: V1/V2 Sites',
            'GVFK Count': metrics['count'],
            'Area (km²)': metrics['area_km2'],
            'Volume (m³)': metrics['volume_m³']
        })

    # Step 5b: Core (substance sites)
    core_metrics = calculate_gvfk_metrics(gvfk_categories['core_gvfks'], gvfk_area_volume)
    rows.append({
        'Step': 'Step 5b: Core (Substance)',
        'GVFK Count': core_metrics['count'],
        'Area (km²)': core_metrics['area_km2'],
        'Volume (m³)': core_metrics['volume_m³']
    })

    # Step 5b⁺: Expanded (substance + branch)
    expanded_metrics = calculate_gvfk_metrics(gvfk_categories['expanded_gvfks'], gvfk_area_volume)
    rows.append({
        'Step': 'Step 5b⁺: Expanded (+Branch)',
        'GVFK Count': expanded_metrics['count'],
        'Area (km²)': expanded_metrics['area_km2'],
        'Volume (m³)': expanded_metrics['volume_m³']
    })

    # Create DataFrame
    df = pd.DataFrame(rows)

    # Add percentage columns
    if baseline_count > 0:
        df['% of Baseline Count'] = (df['GVFK Count'] / baseline_count * 100).round(1)
        df['% of Baseline Area'] = (df['Area (km²)'] / baseline_area * 100).round(1)
        df['% of Baseline Volume'] = (df['Volume (m³)'] / baseline_volume * 100).round(1)

    df['Δ GVFK vs Previous'] = df['GVFK Count'].diff().fillna(0).astype(int)
    df['Δ Area vs Previous (km²)'] = df['Area (km²)'].diff().round(1)
    df['Δ Volume vs Previous (m³)'] = df['Volume (m³)'].diff()

    # Save to CSV
    table_path = os.path.join(output_dir, 'gvfk_progression_table.csv')
    df.to_csv(table_path, index=False)

    print(f"  Saved: gvfk_progression_table.csv")
    print("\n  Progression Summary:")
    print(df.to_string(index=False))

    return df


def create_expansion_breakdown_chart(gvfk_categories, gvfk_area_volume, output_dir):
    """
    Create chart showing contribution of new GVFKs to expanded scenario.

    Args:
        gvfk_categories: Dict with GVFK sets
        gvfk_area_volume: Dict with area/volume data
        output_dir: Output directory path
    """
    print("\n[PHASE 2] Creating expansion breakdown chart...")

    # Calculate metrics
    core_metrics = calculate_gvfk_metrics(gvfk_categories['core_gvfks'], gvfk_area_volume)
    new_metrics = calculate_gvfk_metrics(gvfk_categories['new_gvfks'], gvfk_area_volume)

    # Create subplots
    fig, axes = plt.subplots(1, 3, figsize=(15, 5))

    # Count breakdown
    ax1 = axes[0]
    categories = ['Core\n(Substance)', 'New\n(Branch-only)']
    counts = [core_metrics['count'], new_metrics['count']]
    colors_plot = [COLORS['core'], COLORS['new_gvfk']]
    bars1 = ax1.bar(categories, counts, color=colors_plot, edgecolor='black', linewidth=0.5)
    ax1.set_ylabel('Number of GVFKs', fontweight='bold')
    ax1.set_title('GVFK Count Breakdown', fontweight='bold')
    ax1.grid(axis='y', alpha=0.3, linestyle='--')
    ax1.set_axisbelow(True)
    for bar, count in zip(bars1, counts):
        height = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width()/2., height,
                f'{count:,}', ha='center', va='bottom', fontweight='bold')

    # Area breakdown
    ax2 = axes[1]
    areas = [core_metrics['area_km2'], new_metrics['area_km2']]
    bars2 = ax2.bar(categories, areas, color=colors_plot, edgecolor='black', linewidth=0.5)
    ax2.set_ylabel('Area (km²)', fontweight='bold')
    ax2.set_title('Area Breakdown', fontweight='bold')
    ax2.grid(axis='y', alpha=0.3, linestyle='--')
    ax2.set_axisbelow(True)
    for bar, area in zip(bars2, areas):
        height = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width()/2., height,
                f'{area:,.0f}', ha='center', va='bottom', fontweight='bold')

    # Volume breakdown
    ax3 = axes[2]
    volumes = [core_metrics['volume_m³'], new_metrics['volume_m³']]
    bars3 = ax3.bar(categories, volumes, color=colors_plot, edgecolor='black', linewidth=0.5)
    ax3.set_ylabel('Volume (m³)', fontweight='bold')
    ax3.set_title('Volume Breakdown', fontweight='bold')
    ax3.grid(axis='y', alpha=0.3, linestyle='--')
    ax3.set_axisbelow(True)
    for bar, volume in zip(bars3, volumes):
        height = bar.get_height()
        ax3.text(bar.get_x() + bar.get_width()/2., height,
                f'{volume:,.0f}', ha='center', va='bottom', fontweight='bold')

    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'gvfk_expansion_breakdown.png'), dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  Saved: gvfk_expansion_breakdown.png")

def create_core_branch_area_summary(shapefiles, gvfk_categories, gvfk_area_volume,
                                    substance_sites, branch_sites, new_gvfk_sites, output_dir):
    """Summarise the footprint split between core and branch-only GVFKs."""
    print("[PHASE 2] Creating core vs expanded summary...")

    baseline_metrics = None
    if 'all_dk' in shapefiles:
        baseline_metrics = calculate_gvfk_metrics(set(shapefiles['all_dk']['Navn'].unique()),
                                                  gvfk_area_volume)

    core_metrics = calculate_gvfk_metrics(gvfk_categories['core_gvfks'], gvfk_area_volume)
    new_metrics = calculate_gvfk_metrics(gvfk_categories['new_gvfks'], gvfk_area_volume)
    expanded_metrics = calculate_gvfk_metrics(gvfk_categories['expanded_gvfks'], gvfk_area_volume)

    expanded_site_count = len(substance_sites) + len(branch_sites)

    rows = [
        {
            'Group': 'Baseline (All GVFKs)',
            'GVFK Count': baseline_metrics['count'] if baseline_metrics else None,
            'Sites': None,
            'Area (km²)': baseline_metrics['area_km2'] if baseline_metrics else None,
            'Volume (m³)': baseline_metrics['volume_m³'] if baseline_metrics else None,
            'Avg sites/GVFK': None,
        },
        {
            'Group': 'Core (Step 5b – substance)',
            'GVFK Count': core_metrics['count'],
            'Sites': len(substance_sites),
            'Area (km²)': core_metrics['area_km2'],
            'Volume (m³)': core_metrics['volume_m³'],
            'Avg sites/GVFK': round(len(substance_sites) / core_metrics['count'], 2) if core_metrics['count'] else None,
        },
        {
            'Group': f"Branch-only GVFKs (+{len(gvfk_categories['new_gvfks'])})",
            'GVFK Count': new_metrics['count'],
            'Sites': len(new_gvfk_sites),
            'Area (km²)': new_metrics['area_km2'],
            'Volume (m³)': new_metrics['volume_m³'],
            'Avg sites/GVFK': round(len(new_gvfk_sites) / new_metrics['count'], 2) if new_metrics['count'] else None,
        },
        {
            'Group': 'Expanded (Step 5b⁺ – substance + branch)',
            'GVFK Count': expanded_metrics['count'],
            'Sites': expanded_site_count,
            'Area (km²)': expanded_metrics['area_km2'],
            'Volume (m³)': expanded_metrics['volume_m³'],
            'Avg sites/GVFK': round(expanded_site_count / expanded_metrics['count'], 2) if expanded_metrics['count'] else None,
        },
    ]

    summary_df = pd.DataFrame(rows)

    if baseline_metrics and baseline_metrics['area_km2']:
        summary_df['% of Baseline Area'] = (summary_df['Area (km²)'] / baseline_metrics['area_km2'] * 100).round(1)
    if baseline_metrics and baseline_metrics['volume_m³']:
        summary_df['% of Baseline Volume'] = (summary_df['Volume (m³)'] / baseline_metrics['volume_m³'] * 100).round(1)
    if baseline_metrics and baseline_metrics['count']:
        summary_df['% of Baseline GVFKs'] = (summary_df['GVFK Count'] / baseline_metrics['count'] * 100).round(1)

    summary_df['% of Expanded GVFKs'] = (summary_df['GVFK Count'] / expanded_metrics['count'] * 100).round(1)
    summary_df['% of Expanded Area'] = (summary_df['Area (km²)'] / expanded_metrics['area_km2'] * 100).round(1)
    summary_df['% of Expanded Volume'] = (summary_df['Volume (m³)'] / expanded_metrics['volume_m³'] * 100).round(1)

    out_path = os.path.join(output_dir, 'core_branch_area_summary.csv')
    summary_df.to_csv(out_path, index=False)
    print("  Saved: core_branch_area_summary.csv")
    print(summary_df.to_string(index=False, na_rep='-'))

    return summary_df

def run_phase2(data, output_dir):
    """
    Execute Phase 2: GVFK progression metrics and visualizations.

    Args:
        data: Dictionary from Phase 1 with loaded data
        output_dir: Output directory path
    """
    print("\n" + "=" * 60)
    print("PHASE 2: GVFK PROGRESSION METRICS")
    print("=" * 60)

    gvfk_area_volume = data['gvfk_area_volume']
    gvfk_categories = data['gvfk_categories']
    shapefiles = data['shapefiles']

    # Create visualizations
    create_gvfk_count_progression_chart(shapefiles, gvfk_categories, output_dir)
    create_area_volume_progression_chart(shapefiles, gvfk_categories, gvfk_area_volume, output_dir)
    progression_df = create_progression_table(shapefiles, gvfk_categories, gvfk_area_volume, output_dir)
    create_expansion_breakdown_chart(gvfk_categories, gvfk_area_volume, output_dir)
    create_core_branch_area_summary(shapefiles, gvfk_categories, gvfk_area_volume,
                                    data['substance_sites'], data['branch_sites'], data['sites_in_new_gvfks'],
                                    output_dir)

    print("\n" + "=" * 60)
    print("[OK] PHASE 2 COMPLETE")
    print("=" * 60)

    return progression_df


# ============================================================================
# PHASE 3: THREE-WAY BRANCH/ACTIVITY COMPARISON
# ============================================================================

def count_occurrences(df, column_name):
    """
    Count occurrences from semicolon-separated column.

    Args:
        df: DataFrame with the column
        column_name: Name of column with semicolon-separated values

    Returns:
        pd.Series: Value counts of all occurrences
    """
    all_items = []

    for value in df[column_name].dropna():
        if pd.notna(value) and str(value).strip():
            items = [item.strip() for item in str(value).split(';') if item.strip()]
            all_items.extend(items)

    if all_items:
        return pd.Series(all_items).value_counts()
    else:
        return pd.Series(dtype=int)


def analyze_three_groups(substance_sites, all_branch_sites, new_gvfk_sites):
    """
    Analyze branch/activity patterns across three groups.

    Args:
        substance_sites: Group A - substance sites
        all_branch_sites: Group B - all branch-only sites
        new_gvfk_sites: Group C - sites in new GVFKs only

    Returns:
        dict: Analysis results for branches and activities
    """
    print("\n[PHASE 3] Analyzing three-way comparison...")

    # Count occurrences for each group
    print("  Counting branch occurrences...")
    branches_A = count_occurrences(substance_sites, 'Lokalitetensbranche')
    branches_B = count_occurrences(all_branch_sites, 'Lokalitetensbranche')
    branches_C = count_occurrences(new_gvfk_sites, 'Lokalitetensbranche')

    print("  Counting activity occurrences...")
    activities_A = count_occurrences(substance_sites, 'Lokalitetensaktivitet')
    activities_B = count_occurrences(all_branch_sites, 'Lokalitetensaktivitet')
    activities_C = count_occurrences(new_gvfk_sites, 'Lokalitetensaktivitet')

    # Summary statistics
    print(f"\n  Group A (Substance sites): {len(substance_sites):,} sites")
    print(f"    Branches: {len(branches_A)} unique, {branches_A.sum():,} total occurrences")
    print(f"    Activities: {len(activities_A)} unique, {activities_A.sum():,} total occurrences")

    print(f"\n  Group B (All branch-only): {len(all_branch_sites):,} sites")
    print(f"    Branches: {len(branches_B)} unique, {branches_B.sum():,} total occurrences")
    print(f"    Activities: {len(activities_B)} unique, {activities_B.sum():,} total occurrences")

    print(f"\n  Group C (New GVFK sites only): {len(new_gvfk_sites):,} sites")
    print(f"    Branches: {len(branches_C)} unique, {branches_C.sum():,} total occurrences")
    print(f"    Activities: {len(activities_C)} unique, {activities_C.sum():,} total occurrences")

    # Overlap analysis
    branch_overlap = analyze_overlap(branches_A, branches_B, branches_C, "Branches")
    activity_overlap = analyze_overlap(activities_A, activities_B, activities_C, "Activities")

    return {
        'branches': {'A': branches_A, 'B': branches_B, 'C': branches_C},
        'activities': {'A': activities_A, 'B': activities_B, 'C': activities_C},
        'branch_overlap': branch_overlap,
        'activity_overlap': activity_overlap
    }


def analyze_overlap(series_A, series_B, series_C, label):
    """
    Analyze overlap between three series.

    Args:
        series_A, series_B, series_C: Value count series for each group
        label: Label for printing (Branches/Activities)

    Returns:
        dict: Overlap statistics
    """
    set_A = set(series_A.index)
    set_B = set(series_B.index)
    set_C = set(series_C.index)

    common_all = set_A & set_B & set_C
    common_AC = set_A & set_C
    unique_C = set_C - set_A - set_B

    print(f"\n  {label} Overlap Analysis:")
    print(f"    Unique in A: {len(set_A)}")
    print(f"    Unique in B: {len(set_B)}")
    print(f"    Unique in C: {len(set_C)}")
    print(f"    Common to all three: {len(common_all)}")
    print(f"    Common to A and C: {len(common_AC)}")
    print(f"    Unique to C only: {len(unique_C)}")
    if len(set_C) > 0:
        print(f"    Overlap % (C with A): {len(common_AC)/len(set_C)*100:.1f}%")

    return {
        'unique_A': len(set_A),
        'unique_B': len(set_B),
        'unique_C': len(set_C),
        'common_all': len(common_all),
        'common_AC': len(common_AC),
        'unique_C_only': len(unique_C),
        'overlap_pct': len(common_AC)/len(set_C)*100 if len(set_C) > 0 else 0
    }


def create_three_way_comparison_chart(comparison_data, data_type, output_dir):
    """
    Create three-way comparison horizontal bar chart.

    Args:
        comparison_data: Dict with 'A', 'B', 'C' series
        data_type: 'branches' or 'activities'
        output_dir: Output directory path
    """
    print(f"\n[PHASE 3] Creating three-way {data_type} comparison chart...")

    series_A = comparison_data['A']
    series_B = comparison_data['B']
    series_C = comparison_data['C']

    # Get top 15 by combined total occurrences
    all_items = set(series_A.index) | set(series_B.index) | set(series_C.index)

    item_totals = []
    for item in all_items:
        total = series_A.get(item, 0) + series_B.get(item, 0) + series_C.get(item, 0)
        item_totals.append((item, total))

    item_totals.sort(key=lambda x: x[1], reverse=True)
    top_15_items = [item for item, _ in item_totals[:15]]

    # Get counts for each group
    counts_A = [series_A.get(item, 0) for item in top_15_items]
    counts_B = [series_B.get(item, 0) for item in top_15_items]
    counts_C = [series_C.get(item, 0) for item in top_15_items]

    # Truncate long names
    item_names = [name[:40] + '...' if len(name) > 40 else name for name in top_15_items]

    # Create figure - taller for 15 items
    fig, ax = plt.subplots(figsize=(14, 10))

    y_pos = np.arange(len(item_names))
    width = 0.25

    # Three bars side by side
    bars1 = ax.barh(y_pos - width, counts_A, width,
                    label='Gruppe A: Substanslokalteter (1,740 lokaliteter)',
                    color=COLORS['core'], alpha=0.8)
    bars2 = ax.barh(y_pos, counts_B, width,
                    label='Gruppe B: Alle Branche-kun (3,714 lokaliteter)',
                    color=COLORS['expanded'], alpha=0.8)
    bars3 = ax.barh(y_pos + width, counts_C, width,
                    label='Gruppe C: Kun Nye GVFKs (241 lokaliteter)',
                    color=COLORS['new_gvfk'], alpha=0.8)

    ax.set_yticks(y_pos)
    ax.set_yticklabels(item_names, fontsize=14)
    ax.set_xlabel('Antal Forekomster', fontweight='bold', fontsize=16)
    data_type_da = 'Brancher' if data_type == 'branches' else 'Aktiviteter'
    ax.set_title(f'Top 15 {data_type_da}: Tre-vejs Sammenligning',
                 fontweight='bold', fontsize=20, pad=15)
    ax.legend(fontsize=14, loc='lower right')
    ax.grid(axis='x', alpha=0.3, linestyle='--')
    ax.set_axisbelow(True)

    # Add value labels for Group C (smallest, most interesting)
    for bar, count in zip(bars3, counts_C):
        if count > 0:
            width_val = bar.get_width()
            ax.text(width_val + max(counts_A + counts_B + counts_C)*0.01,
                   bar.get_y() + bar.get_height()/2,
                   f'{int(count)}', ha='left', va='center', fontsize=13,
                   color=COLORS['new_gvfk'], fontweight='bold')

    # Clean styling
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.set_facecolor('#FAFAFA')

    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, f'three_way_{data_type}_comparison.png'),
                dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  Saved: three_way_{data_type}_comparison.png")


def create_two_way_comparison_chart(series_A, series_B, data_type, output_dir):
    """
    Create improved 2-way comparison chart focusing on Group A vs Group B.

    Args:
        series_A: Group A (substance sites) value counts
        series_B: Group B (branch-only sites) value counts
        data_type: 'branches' or 'activities'
        output_dir: Output directory path
    """
    print(f"\n[PHASE 3] Creating improved 2-way {data_type} comparison chart...")

    # Get all unique items
    all_items = set(series_A.index) | set(series_B.index)

    # Calculate combined totals for ranking
    item_totals = []
    for item in all_items:
        total = series_A.get(item, 0) + series_B.get(item, 0)
        item_totals.append((item, total))

    item_totals.sort(key=lambda x: x[1], reverse=True)
    top_20_items = [item for item, _ in item_totals[:20]]

    # Get counts and calculate percentages
    counts_A = [series_A.get(item, 0) for item in top_20_items]
    counts_B = [series_B.get(item, 0) for item in top_20_items]

    total_A = series_A.sum()
    total_B = series_B.sum()

    pct_A = [(c / total_A * 100) if total_A > 0 else 0 for c in counts_A]
    pct_B = [(c / total_B * 100) if total_B > 0 else 0 for c in counts_B]

    # Truncate long names
    item_names = [name[:50] + '...' if len(name) > 50 else name for name in top_20_items]

    # Create figure - larger for better readability
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(22, 12))

    y_pos = np.arange(len(item_names))
    width = 0.35

    # Left panel: Absolute counts
    bars1_abs = ax1.barh(y_pos - width/2, counts_A, width,
                         label='Gruppe A: Substanslokalteter (1,740 lokaliteter)',
                         color='#2E7D32', alpha=0.85)
    bars2_abs = ax1.barh(y_pos + width/2, counts_B, width,
                         label='Gruppe B: Branche-kun lokaliteter (3,714 lokaliteter)',
                         color='#F57C00', alpha=0.85)

    ax1.set_yticks(y_pos)
    ax1.set_yticklabels(item_names, fontsize=14)
    ax1.set_xlabel('Antal Forekomster', fontweight='bold', fontsize=16)
    data_type_da = 'Brancher' if data_type == 'branches' else 'Aktiviteter'
    ax1.set_title(f'Top 20 {data_type_da}: Absolutte Antal',
                  fontweight='bold', fontsize=20, pad=15)
    ax1.legend(fontsize=14, loc='lower right')
    ax1.grid(axis='x', alpha=0.3, linestyle='--')
    ax1.set_axisbelow(True)
    ax1.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)

    # Right panel: Percentage comparison
    bars1_pct = ax2.barh(y_pos - width/2, pct_A, width,
                         label='Gruppe A: Substanslokalteter',
                         color='#2E7D32', alpha=0.85)
    bars2_pct = ax2.barh(y_pos + width/2, pct_B, width,
                         label='Gruppe B: Branche-kun lokaliteter',
                         color='#F57C00', alpha=0.85)

    ax2.set_yticks(y_pos)
    ax2.set_yticklabels([''] * len(item_names))  # No labels on right
    ax2.set_xlabel('Procent af Total (%)', fontweight='bold', fontsize=16)
    data_type_da = 'Brancher' if data_type == 'branches' else 'Aktiviteter'
    ax2.set_title(f'Top 20 {data_type_da}: Relativ Fordeling',
                  fontweight='bold', fontsize=20, pad=15)
    ax2.legend(fontsize=14, loc='lower right')
    ax2.grid(axis='x', alpha=0.3, linestyle='--')
    ax2.set_axisbelow(True)
    ax2.spines['top'].set_visible(False)
    ax2.spines['right'].set_visible(False)

    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, f'two_way_{data_type}_comparison.png'),
                dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  Saved: two_way_{data_type}_comparison.png")


def create_overlap_visualization(branch_overlap, activity_overlap, output_dir):
    """
    Create overlap analysis visualization.

    Args:
        branch_overlap: Branch overlap statistics
        activity_overlap: Activity overlap statistics
        output_dir: Output directory path
    """
    print("\n[PHASE 3] Creating overlap visualization...")

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

    # Branch overlap
    categories = ['Unique\nin A', 'Unique\nin B', 'Unique\nin C',
                  'Common\nto all 3', 'Common\nA & C', 'Unique\nC only']
    values_branch = [
        branch_overlap['unique_A'],
        branch_overlap['unique_B'],
        branch_overlap['unique_C'],
        branch_overlap['common_all'],
        branch_overlap['common_AC'],
        branch_overlap['unique_C_only']
    ]

    colors_plot = ['#1E88E5', '#FFC107', '#FF5722', '#43A047', '#9C27B0', '#E91E63']

    bars1 = ax1.bar(range(len(categories)), values_branch, color=colors_plot, alpha=0.8)
    ax1.set_xticks(range(len(categories)))
    ax1.set_xticklabels(categories, fontsize=13)
    ax1.set_ylabel('Antal', fontweight='bold')
    ax1.set_title('Branche Overlap Analyse', fontweight='bold', fontsize=18)
    ax1.grid(axis='y', alpha=0.3, linestyle='--')

    # Add value labels
    for bar in bars1:
        height = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width()/2., height,
                f'{int(height)}', ha='center', va='bottom', fontsize=14, fontweight='bold')

    # Activity overlap
    values_activity = [
        activity_overlap['unique_A'],
        activity_overlap['unique_B'],
        activity_overlap['unique_C'],
        activity_overlap['common_all'],
        activity_overlap['common_AC'],
        activity_overlap['unique_C_only']
    ]

    bars2 = ax2.bar(range(len(categories)), values_activity, color=colors_plot, alpha=0.8)
    ax2.set_xticks(range(len(categories)))
    ax2.set_xticklabels(categories, fontsize=13)
    ax2.set_ylabel('Antal', fontweight='bold')
    ax2.set_title('Aktivitets Overlap Analyse', fontweight='bold', fontsize=18)
    ax2.grid(axis='y', alpha=0.3, linestyle='--')

    # Add value labels
    for bar in bars2:
        height = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width()/2., height,
                f'{int(height)}', ha='center', va='bottom', fontsize=14, fontweight='bold')

    # Clean styling
    for ax in [ax1, ax2]:
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_facecolor('#FAFAFA')

    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'overlap_analysis.png'),
                dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  Saved: overlap_analysis.png")


def create_comparison_tables(comparison_results, substance_sites, all_branch_sites,
                            new_gvfk_sites, gvfk_categories, output_dir):
    """
    Create comparison tables for three groups.

    Args:
        comparison_results: Results from analyze_three_groups
        substance_sites, all_branch_sites, new_gvfk_sites: Site DataFrames
        gvfk_categories: GVFK categorization dict
        output_dir: Output directory path
    """
    print("\n[PHASE 3] Creating comparison tables...")

    branches_A = comparison_results['branches']['A']
    branches_B = comparison_results['branches']['B']
    branches_C = comparison_results['branches']['C']
    activities_A = comparison_results['activities']['A']
    activities_B = comparison_results['activities']['B']
    activities_C = comparison_results['activities']['C']

    # Table 1: Group characteristics summary
    summary_data = {
        'Metric': [
            'Total sites',
            'GVFKs',
            'Unique branches',
            'Unique activities',
            'Total branch occurrences',
            'Total activity occurrences',
            'Top branch',
            'Top activity',
            'Avg sites per GVFK'
        ],
        'Group A (Substance)': [
            len(substance_sites),
            len(gvfk_categories['core_gvfks']),
            len(branches_A),
            len(activities_A),
            branches_A.sum(),
            activities_A.sum(),
            f"{branches_A.index[0]} ({branches_A.iloc[0]})" if len(branches_A) > 0 else "N/A",
            f"{activities_A.index[0]} ({activities_A.iloc[0]})" if len(activities_A) > 0 else "N/A",
            f"{len(substance_sites) / len(gvfk_categories['core_gvfks']):.1f}"
        ],
        'Group B (All Branch-only)': [
            len(all_branch_sites),
            len(gvfk_categories['branch_gvfks']),
            len(branches_B),
            len(activities_B),
            branches_B.sum(),
            activities_B.sum(),
            f"{branches_B.index[0]} ({branches_B.iloc[0]})" if len(branches_B) > 0 else "N/A",
            f"{activities_B.index[0]} ({activities_B.iloc[0]})" if len(activities_B) > 0 else "N/A",
            f"{len(all_branch_sites) / len(gvfk_categories['branch_gvfks']):.1f}"
        ],
        'Group C (New GVFKs only)': [
            len(new_gvfk_sites),
            len(gvfk_categories['new_gvfks']),
            len(branches_C),
            len(activities_C),
            branches_C.sum(),
            activities_C.sum(),
            f"{branches_C.index[0]} ({branches_C.iloc[0]})" if len(branches_C) > 0 else "N/A",
            f"{activities_C.index[0]} ({activities_C.iloc[0]})" if len(activities_C) > 0 else "N/A",
            f"{len(new_gvfk_sites) / len(gvfk_categories['new_gvfks']):.1f}"
        ]
    }

    summary_df = pd.DataFrame(summary_data)
    summary_path = os.path.join(output_dir, 'three_group_characteristics.csv')
    summary_df.to_csv(summary_path, index=False)
    print(f"  Saved: three_group_characteristics.csv")

    # Table 2: Top 15 branches comparison
    top_branches = create_top_15_comparison_table(branches_A, branches_B, branches_C,
                                                   'branches', output_dir)

    # Table 3: Top 15 activities comparison
    top_activities = create_top_15_comparison_table(activities_A, activities_B, activities_C,
                                                     'activities', output_dir)

    return summary_df, top_branches, top_activities


def create_top_15_comparison_table(series_A, series_B, series_C, data_type, output_dir):
    """
    Create top 15 comparison table.

    Args:
        series_A, series_B, series_C: Value count series for each group
        data_type: 'branches' or 'activities'
        output_dir: Output directory path

    Returns:
        pd.DataFrame: Top 15 comparison table
    """
    # Get top 15 by combined total
    all_items = set(series_A.index) | set(series_B.index) | set(series_C.index)

    item_totals = []
    for item in all_items:
        total = series_A.get(item, 0) + series_B.get(item, 0) + series_C.get(item, 0)
        item_totals.append((
            item,
            series_A.get(item, 0),
            series_B.get(item, 0),
            series_C.get(item, 0),
            total
        ))

    item_totals.sort(key=lambda x: x[4], reverse=True)
    top_15 = item_totals[:15]

    df = pd.DataFrame(top_15, columns=[
        data_type.capitalize()[:-1],  # Branch or Activity (singular)
        'Group A (Substance) occurrences',
        'Group B (All Branch-only) occurrences',
        'Group C (New GVFKs) occurrences',
        'Total occurrences'
    ])

    df.insert(0, 'Rank', range(1, len(df) + 1))

    table_path = os.path.join(output_dir, f'top_15_{data_type}_comparison.csv')
    df.to_csv(table_path, index=False)
    print(f"  Saved: top_15_{data_type}_comparison.csv")

    return df


def run_phase3(data, output_dir):
    """
    Execute Phase 3: Three-way branch/activity comparison.

    Args:
        data: Dictionary from Phase 1 with loaded data
        output_dir: Output directory path
    """
    print("\n" + "=" * 60)
    print("PHASE 3: THREE-WAY BRANCH/ACTIVITY COMPARISON")
    print("=" * 60)

    substance_sites = data['substance_sites']
    all_branch_sites = data['branch_sites']
    new_gvfk_sites = data['sites_in_new_gvfks']
    gvfk_categories = data['gvfk_categories']

    # Analyze three groups
    comparison_results = analyze_three_groups(substance_sites, all_branch_sites, new_gvfk_sites)

    # Create visualizations
    create_three_way_comparison_chart(comparison_results['branches'], 'branches', output_dir)
    create_three_way_comparison_chart(comparison_results['activities'], 'activities', output_dir)

    # Create improved 2-way comparison charts (Group A vs Group B only)
    create_two_way_comparison_chart(comparison_results['branches']['A'],
                                    comparison_results['branches']['B'],
                                    'branches', output_dir)
    create_two_way_comparison_chart(comparison_results['activities']['A'],
                                    comparison_results['activities']['B'],
                                    'activities', output_dir)

    create_overlap_visualization(comparison_results['branch_overlap'],
                                 comparison_results['activity_overlap'],
                                 output_dir)

    # Create tables
    summary_df, top_branches, top_activities = create_comparison_tables(
        comparison_results, substance_sites, all_branch_sites,
        new_gvfk_sites, gvfk_categories, output_dir
    )

    print("\n" + "=" * 60)
    print("[OK] PHASE 3 COMPLETE")
    print("=" * 60)

    return comparison_results


# ============================================================================
# PHASE 4: GEOGRAPHIC VISUALIZATIONS
# ============================================================================

def create_hexagonal_heatmap(sites_gdf, title, output_path, hex_size_km=10):
    """
    Create hexagonal heatmap showing site density.

    Args:
        sites_gdf: GeoDataFrame with site points
        title: Map title
        output_path: Output file path
        hex_size_km: Hexagon size in kilometers
    """
    print(f"\n[PHASE 4] Creating hexagonal heatmap: {title}...")

    if len(sites_gdf) == 0:
        print(f"  Warning: No sites to map")
        return

    # Create figure
    fig, ax = plt.subplots(figsize=(12, 14))

    # Create hexbin (matplotlib's hexagonal binning)
    # Convert hex size from km to map units (assuming EPSG:25832 in meters)
    # Make hexagons 25% larger for better visibility
    hex_size_m = hex_size_km * 2000

    # Get coordinates
    x = sites_gdf.geometry.x
    y = sites_gdf.geometry.y

    # Create hexbin plot with improved color palette
    hexbin = ax.hexbin(x, y, gridsize=int(100000/hex_size_m),
                       cmap='plasma', mincnt=1, alpha=0.8, edgecolors='gray', linewidths=0.2)

    # Add colorbar
    cbar = plt.colorbar(hexbin, ax=ax, label='Number of sites')
    cbar.ax.tick_params(labelsize=14)

    # Set title and labels
    ax.set_title(title, fontsize=20, fontweight='bold', pad=15)
    ax.set_xlabel('Øst (m)', fontsize=16)
    ax.set_ylabel('Nord (m)', fontsize=16)

    # Format axes
    ax.ticklabel_format(style='plain', axis='both')
    ax.grid(True, alpha=0.3, linestyle='--')

    # Add site count annotation
    ax.text(0.02, 0.98, f'Total lokaliteter: {len(sites_gdf):,}',
            transform=ax.transAxes, fontsize=16, fontweight='bold',
            verticalalignment='top',
            bbox=dict(boxstyle='round,pad=0.5', facecolor='white', alpha=0.8))

    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  Saved: {os.path.basename(output_path)}")


def create_side_by_side_heatmaps(substance_detailed, substance_sites, expanded_sites, branch_sites, output_dir):
    """
    Create side-by-side hexagonal heatmaps for Core vs Expanded.

    Args:
        substance_detailed: DataFrame with all site-GVFK-substance combinations (for GVFK counting)
        substance_sites: DataFrame with substance site IDs (deduplicated for spatial plotting)
        expanded_sites: DataFrame with all site IDs (substance + branch)
        branch_sites: DataFrame with branch site IDs
        output_dir: Output directory path
    """
    print("\n[PHASE 4] Creating side-by-side hexagonal heatmaps...")

    # Calculate correct GVFK counts from detailed combinations
    core_gvfk_count = substance_detailed['Closest_GVFK'].dropna().nunique()

    # For expanded scenario, need to combine substance and branch GVFKs
    expanded_gvfks = set(substance_detailed['Closest_GVFK'].dropna().unique())
    if 'Closest_GVFK' in branch_sites.columns:
        expanded_gvfks.update(branch_sites['Closest_GVFK'].dropna().unique())
    expanded_gvfk_count = len(expanded_gvfks)

    # Load Denmark regions as backdrop
    regions_path = INPUT_FILES['denmark_regions_shapefile']
    denmark_gdf = None
    if os.path.exists(regions_path):
        try:
            denmark_gdf = gpd.read_file(regions_path)
            print(f"  Loaded Denmark regions backdrop")
        except:
            print(f"  Warning: Could not load Denmark backdrop")

    # Load site geometries from Step 4 (unique lokalitet distances shapefile)
    # This has ONE geometry per Lokalitet_ID (not duplicates like Step 3)
    step4_sites_path = INPUT_FILES['step4_sites_with_geometry']
    if not os.path.exists(step4_sites_path):
        print(f"  Warning: Step 4 unique lokalitet shapefile not found, skipping heatmaps")
        return

    step4_gdf = gpd.read_file(step4_sites_path)
    print(f"  Loaded {len(step4_gdf)} unique sites from Step 4 with geometries")

    # Step 4 shapefile uses 'Lokalitetsnr' which gets truncated to 'Lokalitets'
    # But our Step 5 CSVs use 'Lokalitet_ID' as the column name
    # Let's check what the ID column is actually called in the shapefile
    if 'Lokalitets' in step4_gdf.columns:
        id_col = 'Lokalitets'
    elif 'Lokalitet_' in step4_gdf.columns:
        id_col = 'Lokalitet_'
    elif 'Lokalitetsn' in step4_gdf.columns:
        id_col = 'Lokalitetsn'
    else:
        print(f"  Warning: Cannot find ID column. Available: {list(step4_gdf.columns)[:10]}")
        return

    # IMPORTANT: Only get sites that are actually in our Step 5 filtered results
    substance_ids = set(substance_sites['Lokalitet_ID'].unique())
    branch_ids = set(expanded_sites['Lokalitet_ID'].unique())

    substance_gdf = step4_gdf[step4_gdf[id_col].isin(substance_ids)].copy()

    # Expanded = substance + branch-only (no duplicates)
    expanded_ids = substance_ids | branch_ids
    expanded_gdf = step4_gdf[step4_gdf[id_col].isin(expanded_ids)].copy()

    print(f"  Matched {len(substance_gdf):,} substance sites with geometries")
    print(f"  Matched {len(expanded_gdf):,} expanded sites with geometries")

    if len(substance_gdf) == 0 or len(expanded_gdf) == 0:
        print(f"  Warning: No sites matched with geometries, skipping heatmaps")
        return

    # Convert to centroids if polygons
    if substance_gdf.geom_type.iloc[0] != 'Point':
        substance_gdf = substance_gdf.copy()
        substance_gdf['geometry'] = substance_gdf.geometry.centroid
    if expanded_gdf.geom_type.iloc[0] != 'Point':
        expanded_gdf = expanded_gdf.copy()
        expanded_gdf['geometry'] = expanded_gdf.geometry.centroid

    # Create side-by-side figure with space for colorbar
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(28, 14))

    # Plot Denmark backdrop on both
    if denmark_gdf is not None:
        denmark_gdf.boundary.plot(ax=ax1, color='black', linewidth=1.5, zorder=1)
        denmark_gdf.plot(ax=ax1, color='white', alpha=0.3, zorder=0)
        denmark_gdf.boundary.plot(ax=ax2, color='black', linewidth=1.5, zorder=1)
        denmark_gdf.plot(ax=ax2, color='white', alpha=0.3, zorder=0)

    # Larger hexagons with better color palette
    gridsize = 27

    # Core scenario (left)
    x1 = substance_gdf.geometry.x
    y1 = substance_gdf.geometry.y
    hexbin1 = ax1.hexbin(x1, y1, gridsize=gridsize,
                         cmap='plasma', mincnt=1, alpha=0.7, edgecolors='none', zorder=2)

    # Expanded scenario (right)
    x2 = expanded_gdf.geometry.x
    y2 = expanded_gdf.geometry.y
    hexbin2 = ax2.hexbin(x2, y2, gridsize=gridsize,
                         cmap='plasma', mincnt=1, alpha=0.7, edgecolors='none', zorder=2)

    # Use same vmin/vmax for shared colorbar
    vmax = max(hexbin1.get_array().max(), hexbin2.get_array().max())
    hexbin1.set_clim(vmin=1, vmax=vmax)
    hexbin2.set_clim(vmin=1, vmax=vmax)

    ax1.set_title(f'Kerne Scenarie: Kun Substanslokalteter\n({len(substance_gdf):,} lokaliteter, {core_gvfk_count} GVFKs)',
                  fontsize=24, fontweight='bold', pad=15)
    ax1.set_aspect('equal')
    ax1.axis('off')

    ax2.set_title(f'Udvidet Scenarie: Substanslokalteter + Branche-kun\n({len(expanded_gdf):,} lokaliteter, {expanded_gvfk_count} GVFKs)',
                  fontsize=24, fontweight='bold', pad=15)
    ax2.set_aspect('equal')
    ax2.axis('off')

    # Match axis limits to Denmark extent
    if denmark_gdf is not None:
        bounds = denmark_gdf.total_bounds
        ax1.set_xlim(bounds[0], bounds[2])
        ax1.set_ylim(bounds[1], bounds[3])
        ax2.set_xlim(bounds[0], bounds[2])
        ax2.set_ylim(bounds[1], bounds[3])
    else:
        # Match to data extent
        all_x = list(x1) + list(x2)
        all_y = list(y1) + list(y2)
        xlim = (min(all_x), max(all_x))
        ylim = (min(all_y), max(all_y))
        ax1.set_xlim(xlim)
        ax1.set_ylim(ylim)
        ax2.set_xlim(xlim)
        ax2.set_ylim(ylim)

    # Add single shared colorbar on the right with proper spacing
    sm = plt.cm.ScalarMappable(cmap='plasma', norm=plt.Normalize(vmin=1, vmax=vmax))
    sm.set_array([])
    cbar = fig.colorbar(sm, ax=[ax1, ax2], shrink=0.8, aspect=30, pad=0.03, fraction=0.046)
    cbar.set_label('Antal lokaliteter', fontsize=18, fontweight='bold')
    cbar.ax.tick_params(labelsize=16)

    plt.tight_layout()
    output_path = os.path.join(output_dir, 'hexagonal_heatmap_comparison.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  Saved: hexagonal_heatmap_comparison.png")


def create_gvfk_choropleth_maps(shapefiles, substance_detailed, substance_sites, all_sites, branch_sites,
                                 gvfk_categories, output_dir):
    """
    Create side-by-side GVFK choropleth maps.

    Args:
        shapefiles: Dict with GVFK shapefiles
        substance_detailed: Core scenario detailed combinations (for accurate GVFK-site counting)
        substance_sites: Core scenario sites (deduplicated, for unique site counts)
        all_sites: Expanded scenario sites (substance + branch, deduplicated)
        branch_sites: Branch scenario sites
        gvfk_categories: GVFK categorization dict
        output_dir: Output directory path
    """
    print("\n[PHASE 4] Creating GVFK choropleth maps...")

    if 'all_dk' not in shapefiles:
        print("  Warning: Denmark shapefile not available")
        return

    all_dk_gdf = shapefiles['all_dk']

    # Count sites per GVFK - use detailed combinations to preserve multi-GVFK relationships
    # For Core: use substance_detailed which has all site-GVFK pairs
    core_counts = substance_detailed.groupby('Closest_GVFK')['Lokalitet_ID'].nunique()

    # For Expanded: combine substance and branch site-GVFK pairs
    expanded_detailed = pd.concat([
        substance_detailed[['Closest_GVFK', 'Lokalitet_ID']],
        branch_sites[['Closest_GVFK', 'Lokalitet_ID']] if 'Closest_GVFK' in branch_sites.columns else pd.DataFrame()
    ], ignore_index=True)
    expanded_counts = expanded_detailed.groupby('Closest_GVFK')['Lokalitet_ID'].nunique()

    # Merge with GeoDataFrame
    core_gdf = all_dk_gdf.copy()
    core_gdf['site_count'] = core_gdf['Navn'].map(core_counts).fillna(0)
    core_gdf = core_gdf[core_gdf['site_count'] > 0]  # Only GVFKs with sites

    expanded_gdf = all_dk_gdf.copy()
    expanded_gdf['site_count'] = expanded_gdf['Navn'].map(expanded_counts).fillna(0)
    expanded_gdf = expanded_gdf[expanded_gdf['site_count'] > 0]

    # Mark new GVFKs
    expanded_gdf['is_new'] = expanded_gdf['Navn'].isin(gvfk_categories['new_gvfks'])

    # Create side-by-side figure - bigger maps with space for colorbar
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(28, 14))

    # Use same vmin/vmax for shared colorbar
    vmin = 1
    vmax = max(core_gdf['site_count'].max(), expanded_gdf['site_count'].max())

    # Calculate dynamic counts for titles
    core_unique_sites = substance_sites['Lokalitet_ID'].nunique()
    expanded_unique_sites = all_sites['Lokalitet_ID'].nunique()
    core_gvfk_count = len(core_gdf)
    expanded_gvfk_count = len(expanded_gdf)
    new_gvfk_count = len(gvfk_categories['new_gvfks'])

    # Core scenario (left)
    core_gdf.plot(column='site_count', ax=ax1, legend=False,
                  cmap='Oranges', edgecolor='black', linewidth=0.3,
                  vmin=vmin, vmax=vmax)
    ax1.set_title(f'Kerne Scenarie: {core_gvfk_count} GVFKs med Substanslokalteter\n({core_unique_sites:,} lokaliteter total)',
                  fontsize=24, fontweight='bold', pad=15)
    ax1.set_aspect('equal')
    ax1.axis('off')

    # Expanded scenario (right)
    shared_expanded = expanded_gdf[~expanded_gdf['is_new']]
    new_expanded = expanded_gdf[expanded_gdf['is_new']]

    # Plot shared GVFKs first
    shared_expanded.plot(column='site_count', ax=ax2, legend=False,
                        cmap='Oranges', edgecolor='black', linewidth=0.3,
                        vmin=vmin, vmax=vmax)

    # Overlay new GVFKs with special styling
    if len(new_expanded) > 0:
        new_expanded.plot(ax=ax2, facecolor='none', edgecolor='red',
                         linewidth=2.5, hatch='///', alpha=0.7)

    ax2.set_title(f'Udvidet Scenarie: {expanded_gvfk_count} GVFKs ({core_gvfk_count} kerne + {new_gvfk_count} nye)\n({expanded_unique_sites:,} lokaliteter total)\nRød skraveret = Nye GVFKs fra branche-kun lokaliteter',
                  fontsize=24, fontweight='bold', pad=15)
    ax2.set_aspect('equal')
    ax2.axis('off')

    # Add single shared colorbar on the right with proper spacing
    sm = plt.cm.ScalarMappable(cmap='Oranges', norm=plt.Normalize(vmin=vmin, vmax=vmax))
    sm.set_array([])
    cbar = fig.colorbar(sm, ax=[ax1, ax2], shrink=0.8, aspect=30, pad=0.03, fraction=0.046)
    cbar.set_label('Lokaliteter per GVFK', fontsize=18, fontweight='bold')
    cbar.ax.tick_params(labelsize=16)

    plt.tight_layout()
    output_path = os.path.join(output_dir, 'gvfk_choropleth_comparison.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  Saved: gvfk_choropleth_comparison.png")


def create_regional_distribution_table(substance_sites, all_sites, gvfk_categories, output_dir):
    """
    Create regional distribution comparison table.

    Args:
        substance_sites: Core scenario sites
        all_sites: Expanded scenario sites
        gvfk_categories: GVFK categorization
        output_dir: Output directory path
    """
    print("\n[PHASE 4] Creating regional distribution table...")

    if 'Regionsnavn' not in substance_sites.columns:
        print("  Warning: No region information available")
        return

    # Count by region for each scenario
    core_region_sites = substance_sites.groupby('Regionsnavn').size()
    core_region_gvfks = substance_sites.groupby('Regionsnavn')['Closest_GVFK'].nunique()

    expanded_region_sites = all_sites.groupby('Regionsnavn').size()
    expanded_region_gvfks = all_sites.groupby('Regionsnavn')['Closest_GVFK'].nunique()

    # Build table
    regions = sorted(set(core_region_sites.index) | set(expanded_region_sites.index))

    table_data = []
    for region in regions:
        core_s = core_region_sites.get(region, 0)
        core_g = core_region_gvfks.get(region, 0)
        exp_s = expanded_region_sites.get(region, 0)
        exp_g = expanded_region_gvfks.get(region, 0)

        table_data.append({
            'Region': region,
            'Core GVFKs': core_g,
            'Core Sites': core_s,
            'Expanded GVFKs': exp_g,
            'Expanded Sites': exp_s,
            'Change GVFKs': exp_g - core_g,
            'Change Sites': exp_s - core_s
        })

    # Add total row
    table_data.append({
        'Region': 'TOTAL',
        'Core GVFKs': len(gvfk_categories['core_gvfks']),
        'Core Sites': len(substance_sites),
        'Expanded GVFKs': len(gvfk_categories['expanded_gvfks']),
        'Expanded Sites': len(all_sites),
        'Change GVFKs': len(gvfk_categories['new_gvfks']),
        'Change Sites': len(all_sites) - len(substance_sites)
    })

    df = pd.DataFrame(table_data)
    table_path = os.path.join(output_dir, 'regional_distribution.csv')
    df.to_csv(table_path, index=False)

    print(f"  Saved: regional_distribution.csv")
    print("\n  Regional Distribution:")
    print(df.to_string(index=False))

    return df


def create_new_gvfk_characteristics_table(new_gvfk_sites, gvfk_area_volume,
                                          gvfk_categories, output_dir):
    """
    Create detailed characteristics table for the 92 new GVFKs.

    Args:
        new_gvfk_sites: Sites in new GVFKs
        gvfk_area_volume: Area/volume lookup dict
        gvfk_categories: GVFK categorization
        output_dir: Output directory path
    """
    print("\n[PHASE 4] Creating new GVFK characteristics table...")

    new_gvfks = sorted(gvfk_categories['new_gvfks'])

    table_data = []
    for gvfk in new_gvfks:
        sites_in_gvfk = new_gvfk_sites[new_gvfk_sites['Closest_GVFK'] == gvfk]

        # Get area/volume
        area = gvfk_area_volume.get(gvfk, {}).get('area_km2', 0)
        volume = gvfk_area_volume.get(gvfk, {}).get('volume_m³', 0)

        # Get top branch and activity
        branches = count_occurrences(sites_in_gvfk, 'Lokalitetensbranche')
        activities = count_occurrences(sites_in_gvfk, 'Lokalitetensaktivitet')

        top_branch = f"{branches.index[0]} ({branches.iloc[0]})" if len(branches) > 0 else "N/A"
        top_activity = f"{activities.index[0]} ({activities.iloc[0]})" if len(activities) > 0 else "N/A"

        table_data.append({
            'GVFK': gvfk,
            'Sites': len(sites_in_gvfk),
            'Area (km²)': f"{area:.2f}",
            'Volume (m³)': f"{volume:.0f}",
            'Top Branch': top_branch,
            'Top Activity': top_activity
        })

    df = pd.DataFrame(table_data)
    table_path = os.path.join(output_dir, 'new_gvfk_characteristics.csv')
    df.to_csv(table_path, index=False)

    print(f"  Saved: new_gvfk_characteristics.csv")
    print(f"  Characterized {len(new_gvfks)} new GVFKs")

    return df


def run_phase4(data, output_dir):
    """
    Execute Phase 4: Geographic visualizations.

    Args:
        data: Dictionary from Phase 1 with loaded data
        output_dir: Output directory path
    """
    print("\n" + "=" * 60)
    print("PHASE 4: GEOGRAPHIC VISUALIZATIONS")
    print("=" * 60)

    substance_detailed = data['substance_detailed']
    substance_sites = data['substance_sites']
    branch_sites = data['branch_sites']
    new_gvfk_sites = data['sites_in_new_gvfks']
    gvfk_categories = data['gvfk_categories']
    gvfk_area_volume = data['gvfk_area_volume']
    shapefiles = data['shapefiles']

    # Combine for expanded scenario
    all_sites = pd.concat([substance_sites, branch_sites], ignore_index=True)

    # Create hexagonal heatmaps
    create_side_by_side_heatmaps(substance_detailed, substance_sites, all_sites, branch_sites, output_dir)

    # Create choropleth maps
    create_gvfk_choropleth_maps(shapefiles, substance_detailed, substance_sites, all_sites, branch_sites,
                                gvfk_categories, output_dir)

    # Create tables
    regional_df = create_regional_distribution_table(substance_sites, all_sites,
                                                     gvfk_categories, output_dir)

    new_gvfk_df = create_new_gvfk_characteristics_table(new_gvfk_sites, gvfk_area_volume,
                                                        gvfk_categories, output_dir)

    print("\n" + "=" * 60)
    print("[OK] PHASE 4 COMPLETE")
    print("=" * 60)

    return {
        'regional_distribution': regional_df,
        'new_gvfk_characteristics': new_gvfk_df
    }



# ============================================================================
# MAIN EXECUTION
# ============================================================================

def compare_core_vs_new_gvfks(gvfk_categories, gvfk_area_volume, output_dir):
    """
    Compare characteristics of Core 218 GVFKs vs New 92 GVFKs.

    Args:
        gvfk_categories: GVFK categorization dict
        gvfk_area_volume: Area/volume lookup dict
        output_dir: Output directory path
    """
    print("\n" + "=" * 60)
    print("PHASE 5: CORE vs NEW GVFK COMPARISON")
    print("=" * 60)

    core_gvfks = list(gvfk_categories['core_gvfks'])
    new_gvfks = list(gvfk_categories['new_gvfks'])

    # Collect area and volume data
    core_areas = []
    core_volumes = []
    new_areas = []
    new_volumes = []

    for gvfk in core_gvfks:
        if gvfk in gvfk_area_volume:
            core_areas.append(gvfk_area_volume[gvfk]['area_km2'])
            core_volumes.append(gvfk_area_volume[gvfk]['volume_m³'])

    for gvfk in new_gvfks:
        if gvfk in gvfk_area_volume:
            new_areas.append(gvfk_area_volume[gvfk]['area_km2'])
            new_volumes.append(gvfk_area_volume[gvfk]['volume_m³'])

    # Load V1/V2 site counts from Step 3
    print("\n[PHASE 5] Loading V1/V2 site data from Step 3...")
    step3_path = INPUT_FILES['step3_sites_shapefile']
    core_v1v2_counts = {}
    new_v1v2_counts = {}

    if os.path.exists(step3_path):
        import geopandas as gpd
        step3_gdf = gpd.read_file(step3_path)

        # Check what column contains the GVFK name and site ID
        gvfk_col = None
        site_id_col = None

        for col in ['Closest_GVFK', 'Navn', 'GVFK', 'gvfk_navn']:
            if col in step3_gdf.columns:
                gvfk_col = col
                break

        for col in ['Lokalitet_', 'Lokalitetsnr', 'Lokalitet_ID', 'site_id']:
            if col in step3_gdf.columns:
                site_id_col = col
                break

        if gvfk_col and site_id_col:
            # Count UNIQUE sites per GVFK (not total rows which include duplicates)
            v1v2_per_gvfk = step3_gdf.groupby(gvfk_col)[site_id_col].nunique()

            for gvfk in core_gvfks:
                core_v1v2_counts[gvfk] = v1v2_per_gvfk.get(gvfk, 0)

            for gvfk in new_gvfks:
                new_v1v2_counts[gvfk] = v1v2_per_gvfk.get(gvfk, 0)

            total_unique_sites = step3_gdf[site_id_col].nunique()
            print(f"  Loaded V1/V2 site data: {total_unique_sites:,} unique sites ({len(step3_gdf):,} total rows with duplicates)")

            # Store total unique sites in each category (for summary table later)
            core_v1v2_counts['_total_unique'] = step3_gdf[step3_gdf[gvfk_col].isin(core_gvfks)][site_id_col].nunique()
            new_v1v2_counts['_total_unique'] = step3_gdf[step3_gdf[gvfk_col].isin(new_gvfks)][site_id_col].nunique()
        else:
            print(f"  Warning: Could not find GVFK column in Step 3 data. Available columns: {list(step3_gdf.columns)[:10]}")
    else:
        print(f"  Warning: Step 3 data not found at {step3_path}, V1/V2 counts unavailable")

    # Extract total unique counts (stored with special key)
    core_total_v1v2 = core_v1v2_counts.pop('_total_unique', 0)
    new_total_v1v2 = new_v1v2_counts.pop('_total_unique', 0)

    # Convert to lists (excluding the special _total_unique key which we already popped)
    core_v1v2 = list(core_v1v2_counts.values())
    new_v1v2 = list(new_v1v2_counts.values())

    # Create summary statistics
    summary_data = {
        'Metric': ['Count', 'Mean Area (km²)', 'Median Area (km²)', 'Min Area (km²)',
                   'Max Area (km²)', 'Mean Volume (m³)', 'Median Volume (m³)',
                   'Mean V1/V2 Sites', 'Median V1/V2 Sites', 'Total V1/V2 Sites (unique)'],
        'Core 218 GVFKs': [
            len(core_gvfks),
            np.mean(core_areas) if core_areas else 0,
            np.median(core_areas) if core_areas else 0,
            np.min(core_areas) if core_areas else 0,
            np.max(core_areas) if core_areas else 0,
            np.mean(core_volumes) if core_volumes else 0,
            np.median(core_volumes) if core_volumes else 0,
            np.mean(core_v1v2) if core_v1v2 else 0,
            np.median(core_v1v2) if core_v1v2 else 0,
            core_total_v1v2
        ],
        'New 92 GVFKs': [
            len(new_gvfks),
            np.mean(new_areas) if new_areas else 0,
            np.median(new_areas) if new_areas else 0,
            np.min(new_areas) if new_areas else 0,
            np.max(new_areas) if new_areas else 0,
            np.mean(new_volumes) if new_volumes else 0,
            np.median(new_volumes) if new_volumes else 0,
            np.mean(new_v1v2) if new_v1v2 else 0,
            np.median(new_v1v2) if new_v1v2 else 0,
            new_total_v1v2
        ]
    }

    summary_df = pd.DataFrame(summary_data)
    summary_path = os.path.join(output_dir, 'core_vs_new_gvfk_comparison.csv')
    summary_df.to_csv(summary_path, index=False)

    print(f"\n  Saved: core_vs_new_gvfk_comparison.csv")
    print("\n  Summary Statistics:")
    print(summary_df.to_string(index=False))

    # Create box plot comparison
    print("\n[PHASE 5] Creating comparison visualizations...")

    fig, axes = plt.subplots(2, 2, figsize=(16, 12))

    # Area box plot
    ax1 = axes[0, 0]
    bp1 = ax1.boxplot([core_areas, new_areas], labels=['Kerne 218 GVFKs', 'Nye 92 GVFKs'],
                       patch_artist=True, showmeans=True)
    bp1['boxes'][0].set_facecolor('#4CAF50')
    bp1['boxes'][1].set_facecolor('#FFC107')
    ax1.set_ylabel('Areal (km²)', fontsize=16, fontweight='bold')
    ax1.set_title('GVFK Areal Fordeling', fontsize=20, fontweight='bold')
    ax1.grid(axis='y', alpha=0.3)

    # Volume box plot
    ax2 = axes[0, 1]
    bp2 = ax2.boxplot([core_volumes, new_volumes], labels=['Kerne 218 GVFKs', 'Nye 92 GVFKs'],
                       patch_artist=True, showmeans=True)
    bp2['boxes'][0].set_facecolor('#4CAF50')
    bp2['boxes'][1].set_facecolor('#FFC107')
    ax2.set_ylabel('Volumen (m³)', fontsize=16, fontweight='bold')
    ax2.set_title('GVFK Volumen Fordeling', fontsize=20, fontweight='bold')
    ax2.grid(axis='y', alpha=0.3)
    ax2.ticklabel_format(style='scientific', axis='y', scilimits=(0,0))

    # Risk assessment site counts per GVFK
    ax3 = axes[1, 0]

    # Load actual risk assessment site data
    core_detailed_path = INPUT_FILES['step5_compound_detailed']
    branch_sites_path = INPUT_FILES['step5_parked_sites']

    core_site_counts = []
    new_site_counts = []

    if os.path.exists(core_detailed_path):
        core_detailed_df = pd.read_csv(core_detailed_path)

        # Standardize column names
        if 'GVFK' in core_detailed_df.columns and 'Closest_GVFK' not in core_detailed_df.columns:
            core_detailed_df = core_detailed_df.rename(columns={'GVFK': 'Closest_GVFK'})

        if 'Closest_GVFK' in core_detailed_df.columns:
            # Count unique sites per GVFK (deduplicate site-GVFK combinations by site within each GVFK)
            core_sites_per_gvfk = core_detailed_df.groupby('Closest_GVFK')['Lokalitet_ID'].nunique()
            for gvfk in core_gvfks:
                core_site_counts.append(core_sites_per_gvfk.get(gvfk, 0))

    if os.path.exists(branch_sites_path):
        branch_df = pd.read_csv(branch_sites_path)

        # Standardize column names
        if 'GVFK' in branch_df.columns and 'Closest_GVFK' not in branch_df.columns:
            branch_df = branch_df.rename(columns={'GVFK': 'Closest_GVFK'})
        if 'Distance_to_River_m' in branch_df.columns and 'Final_Distance_m' not in branch_df.columns:
            branch_df = branch_df.rename(columns={'Distance_to_River_m': 'Final_Distance_m'})

        # Filter to <=500m and exclude losseplads
        if 'Final_Distance_m' in branch_df.columns:
            branch_df = branch_df[branch_df['Final_Distance_m'] <= 500]
        if 'Losseplads_Flag' in branch_df.columns:
            branch_df = branch_df[branch_df['Losseplads_Flag'] == False]

        if 'Closest_GVFK' in branch_df.columns:
            # Count sites per GVFK for new 92
            branch_sites_per_gvfk = branch_df.groupby('Closest_GVFK').size()
            for gvfk in new_gvfks:
                new_site_counts.append(branch_sites_per_gvfk.get(gvfk, 0))

    if core_site_counts and new_site_counts:
        bp3 = ax3.boxplot([core_site_counts, new_site_counts],
                           labels=['Kerne 218 GVFKs\n(1,743 total lokaliteter)', 'Nye 92 GVFKs\n(~241 total lokaliteter)'],
                           patch_artist=True, showmeans=True)
        bp3['boxes'][0].set_facecolor('#4CAF50')
        bp3['boxes'][1].set_facecolor('#FFC107')
        ax3.set_ylabel('Risikovurderingslokalteter per GVFK', fontsize=16, fontweight='bold')
        ax3.set_title('Lokalitetsantal Fordeling', fontsize=20, fontweight='bold')
        ax3.grid(axis='y', alpha=0.3)
    else:
        ax3.text(0.5, 0.5, 'Lokalitetsdata ikke tilgængelig', ha='center', va='center',
                 fontsize=16, transform=ax3.transAxes)
        ax3.set_title('Lokalitetsantal Fordeling', fontsize=20, fontweight='bold')

    # Volume histogram overlay
    ax4 = axes[1, 1]
    ax4.hist(core_volumes, bins=30, alpha=0.6, label='Kerne 218 GVFKs', color='#4CAF50', edgecolor='black')
    ax4.hist(new_volumes, bins=30, alpha=0.6, label='Nye 92 GVFKs', color='#FFC107', edgecolor='black')
    ax4.set_xlabel('Volumen (m³)', fontsize=16, fontweight='bold')
    ax4.set_ylabel('Frekvens', fontsize=16, fontweight='bold')
    ax4.set_title('Volumen Fordelings Histogram', fontsize=20, fontweight='bold')
    ax4.legend(fontsize=14)
    ax4.grid(axis='y', alpha=0.3)
    ax4.ticklabel_format(style='scientific', axis='x', scilimits=(0,0))

    plt.tight_layout()
    comparison_path = os.path.join(output_dir, 'core_vs_new_gvfk_distributions.png')
    plt.savefig(comparison_path, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"  Saved: core_vs_new_gvfk_distributions.png")

    print("\n" + "=" * 60)
    print("[OK] PHASE 5 COMPLETE")
    print("=" * 60)

    return summary_df


def export_tables_to_excel(output_dir):
    """
    Convert all CSV tables to nicely formatted Excel files.

    Args:
        output_dir: Directory containing CSV files
    """
    print("\n" + "=" * 60)
    print("PHASE 6: EXPORTING FORMATTED EXCEL TABLES")
    print("=" * 60)

    import glob
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Find all CSV files
    csv_files = glob.glob(os.path.join(output_dir, '*.csv'))

    if not csv_files:
        print("  No CSV files found to export")
        return

    print(f"\n  Found {len(csv_files)} CSV files to convert")

    for csv_path in csv_files:
        # Read CSV
        df = pd.read_csv(csv_path)

        # Create Excel filename
        excel_path = csv_path.replace('.csv', '_formatted.xlsx')

        # Write to Excel
        df.to_excel(excel_path, index=False, sheet_name='Data')

        # Load workbook for formatting
        wb = load_workbook(excel_path)
        ws = wb.active

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        cell_alignment = Alignment(horizontal="left", vertical="center")
        number_alignment = Alignment(horizontal="right", vertical="center")

        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Format data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

                # Align numbers to the right, text to the left
                if isinstance(cell.value, (int, float)):
                    cell.alignment = number_alignment
                    # Format large numbers with commas
                    if isinstance(cell.value, int) and cell.value > 999:
                        cell.number_format = '#,##0'
                    elif isinstance(cell.value, float):
                        # Scientific notation for very large numbers
                        if abs(cell.value) > 1e6:
                            cell.number_format = '0.00E+00'
                        else:
                            cell.number_format = '#,##0.0'
                else:
                    cell.alignment = cell_alignment

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass

            # Set width with some padding, max 50 characters
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Set row height for header
        ws.row_dimensions[1].height = 30

        # Save formatted workbook
        wb.save(excel_path)

        print(f"  Exported: {os.path.basename(excel_path)}")

    print("\n" + "=" * 60)
    print("[OK] PHASE 6 COMPLETE")
    print("=" * 60)
    print(f"\n  All tables exported to Excel with formatting")


def run_step6_analysis():
    """
    Run complete Step 6 final analysis.
    """
    print("\n" + "=" * 80)
    print("STEP 6: FINAL COMPREHENSIVE ANALYSIS")
    print("=" * 80)
    print("\nComparing Core vs Expanded scenarios:")
    print("  Core: Step 5b compound-specific sites (substance data)")
    print("  Expanded: Core + branch-only sites (<=500m, no losseplads)")
    print("=" * 80)

    # Create output directory
    output_dir = get_visualization_path('risikovurdering', 'step6')
    os.makedirs(output_dir, exist_ok=True)
    print(f"\nOutput directory: {output_dir}")

    # Phase 1: Load data
    data = run_phase1()

    # Phase 2: GVFK progression metrics
    progression_df = run_phase2(data, output_dir)

    # Phase 3: Three-way branch/activity comparison
    comparison_results = run_phase3(data, output_dir)

    # Phase 4: Geographic visualizations
    geographic_results = run_phase4(data, output_dir)

    # Phase 5: Core vs New GVFK comparison
    gvfk_comparison = compare_core_vs_new_gvfks(
        data['gvfk_categories'],
        data['gvfk_area_volume'],
        output_dir
    )

    # Phase 6: Export formatted Excel tables
    export_tables_to_excel(output_dir)

    print("\n" + "=" * 80)
    print("[OK] STEP 6 ANALYSIS COMPLETE (ALL PHASES)")
    print("=" * 80)
    print(f"\nResults saved to: {output_dir}")
    print(f"\nGenerated outputs:")
    print(f"  - 4 progression visualizations")
    print(f"  - 4 comparison charts")
    print(f"  - 2 geographic maps")
    print(f"  - 7 CSV tables + 7 formatted Excel tables")
    print(f"  - 1 GVFK distribution comparison")


if __name__ == "__main__":
    run_step6_analysis()
